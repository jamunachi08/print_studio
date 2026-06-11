# Copyright (c) 2026, Neotec Integrated Solution
# Reads an existing format (PDF / Excel / Word / image) into a Print Studio
# layout draft. Static text & headings are placed; the user then converts the
# ones that should be dynamic into variables (fields) inside the designer.
import base64
import io

import frappe
from frappe import _

PT2MM = 25.4 / 72.0
A4_W = 210.0
MAX_ELEMENTS = 1200


def _txt(x, y, w, h, text, size=9, bold=False, italic=False, align="left", color="#1f2733"):
    h = max(h, size * 0.42)          # keep boxes tight to one line to reduce overlap
    return {
        "type": "text",
        "x": round(x, 1), "y": round(y, 1), "w": round(max(2, w), 1), "h": round(max(3, h), 1),
        "props": {"text": text, "fontSize": round(size, 1), "bold": bool(bold), "italic": bool(italic),
                  "align": align or "left", "color": color, "bg": "", "border": "none", "nowrap": True},
    }


def _line(x, y, w, h, vertical=False):
    if vertical:
        return {"type": "vline", "x": round(x, 1), "y": round(y, 1), "w": 0.3, "h": round(max(1, h), 1),
                "props": {"color": "#333", "thick": 0.3}}
    return {"type": "hline", "x": round(x, 1), "y": round(y, 1), "w": round(max(1, w), 1), "h": 0.3,
            "props": {"color": "#333", "thick": 0.3}}


def _rect(x, y, w, h):
    return {"type": "rect", "x": round(x, 1), "y": round(y, 1), "w": round(max(2, w), 1), "h": round(max(2, h), 1),
            "props": {"border": "all", "borderColor": "#333", "borderW": 0.3, "bg": ""}}


def _wrap(doctype, elements, page_h):
    return {
        "doctype": doctype,
        "page": {"size": "A4", "orientation": "portrait",
                 "marginTop": 6, "marginBottom": 6, "marginLeft": 6, "marginRight": 6},
        "bands": [{"type": "report_header", "height": round(max(40, page_h), 1),
                   "showOn": "first_page", "elements": elements[:MAX_ELEMENTS]}],
        "__imported": 1,
    }


# ----------------------------------------------------------------- PDF
def parse_pdf(data, doctype):
    try:
        import fitz  # PyMuPDF
    except Exception:
        frappe.throw(_("PDF import needs PyMuPDF. Install it on the bench: <code>./env/bin/pip install pymupdf</code>"))
    doc = fitz.open(stream=data, filetype="pdf")
    pg = doc[0]
    W = pg.rect.width * PT2MM
    H = pg.rect.height * PT2MM
    sx = A4_W / W if W else 1.0          # scale to A4 width
    els = []
    info = pg.get_text("dict")
    for block in info.get("blocks", []):
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            text = "".join(s.get("text", "") for s in spans).strip()
            if not text:
                continue
            x0, y0, x1, y1 = line["bbox"]
            s0 = spans[0]
            size = s0.get("size", 9)
            bold = bool(s0.get("flags", 0) & 2 ** 4) or "Bold" in (s0.get("font", "") or "")
            italic = bool(s0.get("flags", 0) & 2 ** 1)
            els.append(_txt(x0 * PT2MM * sx, y0 * PT2MM * sx,
                            (x1 - x0) * PT2MM * sx + 2, (y1 - y0) * PT2MM * sx,
                            text, size * sx, bold, italic))
    # lines & boxes
    try:
        for dr in pg.get_drawings()[:400]:
            for it in dr.get("items", []):
                if it[0] == "l":
                    p1, p2 = it[1], it[2]
                    x = min(p1.x, p2.x) * PT2MM * sx
                    y = min(p1.y, p2.y) * PT2MM * sx
                    if abs(p1.y - p2.y) < abs(p1.x - p2.x):
                        els.append(_line(x, y, abs(p2.x - p1.x) * PT2MM * sx, 0))
                    else:
                        els.append(_line(x, y, 0, abs(p2.y - p1.y) * PT2MM * sx, vertical=True))
                elif it[0] == "re":
                    r = it[1]
                    rw = (r.x1 - r.x0) * PT2MM * sx
                    rh = (r.y1 - r.y0) * PT2MM * sx
                    page_area = (W * sx) * (H * sx)
                    # skip the outer page-border box that would cover everything
                    if rw * rh > 0.75 * page_area:
                        continue
                    els.append(_rect(r.x0 * PT2MM * sx, r.y0 * PT2MM * sx, rw, rh))
    except Exception:
        pass
    return _wrap(doctype, els, H * sx)


# ----------------------------------------------------------------- Excel
def parse_xlsx(data, doctype):
    try:
        import openpyxl
    except Exception:
        frappe.throw(_("Excel import needs openpyxl."))
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    COL_MM = 2.05          # approx mm per Excel column-width unit
    DEF_COL = 8.43
    maxc, maxr = ws.max_column, ws.max_row
    xs = [0.0]
    for c in range(1, maxc + 1):
        letter = openpyxl.utils.get_column_letter(c)
        w = ws.column_dimensions[letter].width or DEF_COL
        xs.append(xs[-1] + w * COL_MM)
    ys = [0.0]
    for r in range(1, maxr + 1):
        h = ws.row_dimensions[r].height or 15.0     # points
        ys.append(ys[-1] + h * PT2MM)

    merged_start = {}
    covered = set()
    for rng in ws.merged_cells.ranges:
        merged_start[(rng.min_row, rng.min_col)] = (rng.max_row, rng.max_col)
        for rr in range(rng.min_row, rng.max_row + 1):
            for cc in range(rng.min_col, rng.max_col + 1):
                if (rr, cc) != (rng.min_row, rng.min_col):
                    covered.add((rr, cc))

    els = []
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(r, c)
            v = cell.value
            if v in (None, ""):
                continue
            x, y = xs[c - 1], ys[r - 1]
            if (r, c) in merged_start:
                mr, mc = merged_start[(r, c)]
                w = xs[mc] - xs[c - 1]
                hh = ys[mr] - ys[r - 1]
            else:
                w = xs[c] - xs[c - 1]
                hh = ys[r] - ys[r - 1]
            f = cell.font
            al = (cell.alignment.horizontal or "left")
            if al == "general":
                al = "right" if isinstance(v, (int, float)) else "left"
            els.append(_txt(x, y, w, hh, str(v), (f.size or 10), bool(f.bold), bool(f.italic), al))
    return _wrap(doctype, els, ys[-1])


# ----------------------------------------------------------------- Word
def parse_docx(data, doctype):
    try:
        import docx
        from docx.oxml.ns import qn
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except Exception:
        frappe.throw(_("Word import needs python-docx. Install: <code>./env/bin/pip install python-docx</code>"))
    d = docx.Document(io.BytesIO(data))
    els = []
    y = 6.0
    for child in d.element.body.iterchildren():
        if child.tag == qn("w:p"):
            para = Paragraph(child, d)
            text = (para.text or "").strip()
            if not text:
                y += 3
                continue
            size = 11
            bold = False
            if para.runs:
                r0 = para.runs[0]
                size = (r0.font.size.pt if r0.font.size else None) or (
                    18 if (para.style and "Heading" in (para.style.name or "")) else 10)
                bold = bool(r0.bold) or ("Heading" in (para.style.name or "") if para.style else False)
            al = {0: "left", 1: "center", 2: "right", 3: "left"}.get(
                int(para.alignment) if para.alignment is not None else 0, "left")
            h = size * 0.5
            els.append(_txt(8, y, 194, h, text, size, bold, False, al))
            y += h + 1.5
        elif child.tag == qn("w:tbl"):
            tbl = Table(child, d)
            ncol = len(tbl.columns)
            cols = []
            head = tbl.rows[0].cells if tbl.rows else []
            for i in range(ncol):
                label = (head[i].text.strip() if i < len(head) else "Col {0}".format(i + 1)) or "Col {0}".format(i + 1)
                cols.append({"field": "", "label": label, "w": round(100 / max(1, ncol)), "align": "left"})
            th = max(10, len(tbl.rows) * 6)
            els.append({"type": "table", "x": 8, "y": round(y, 1), "w": 194, "h": th,
                        "props": {"source": "", "columns": cols, "fontSize": 8, "headerBg": "#eef4f7"}})
            y += th + 3
    return _wrap(doctype, els, y + 6)


# ----------------------------------------------------------------- Image (tracing layer)
def parse_image(data, doctype, ext):
    from PIL import Image
    im = Image.open(io.BytesIO(data))
    if im.mode not in ("RGB", "RGBA"):
        im = im.convert("RGB")
    # downscale to keep the data URI small
    if im.width > 1240:
        im = im.resize((1240, int(im.height * 1240 / im.width)))
    buf = io.BytesIO()
    im.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode()
    Wmm = A4_W
    Hmm = Wmm * im.height / im.width
    bg = {"type": "image", "x": 0, "y": 0, "w": round(Wmm, 1), "h": round(Hmm, 1),
          "props": {"src": "data:image/png;base64," + b64, "fit": "contain"}}
    els = [bg]
    # optional OCR if the server has it (never required)
    try:
        import pytesseract
        from PIL import Image as _I
        data_ocr = pytesseract.image_to_data(im, output_type=pytesseract.Output.DICT)
        sx = Wmm / im.width
        n = len(data_ocr["text"])
        for i in range(n):
            t = (data_ocr["text"][i] or "").strip()
            if not t or int(data_ocr["conf"][i]) < 60:
                continue
            els.append(_txt(data_ocr["left"][i] * sx, data_ocr["top"][i] * sx,
                            data_ocr["width"][i] * sx + 2, data_ocr["height"][i] * sx,
                            t, max(7, data_ocr["height"][i] * sx * 0.7)))
    except Exception:
        pass
    return _wrap(doctype, els, Hmm)


# ----------------------------------------------------------------- dispatch
def parse_pdf_image(data, doctype):
    """Rasterise the PDF page to an exact background image (faithful template)."""
    try:
        import fitz
    except Exception:
        frappe.throw(_("PDF import needs PyMuPDF: <code>./env/bin/pip install pymupdf</code>"))
    doc = fitz.open(stream=data, filetype="pdf")
    pg = doc[0]
    pix = pg.get_pixmap(matrix=fitz.Matrix(2, 2))      # ~144 dpi
    png = pix.tobytes("png")
    b64 = base64.b64encode(png).decode()
    Wmm = pg.rect.width * PT2MM
    Hmm = pg.rect.height * PT2MM
    landscape = Wmm > Hmm
    page_w = 297.0 if landscape else 210.0
    scale = page_w / Wmm if Wmm else 1.0
    iw, ih = Wmm * scale, Hmm * scale
    bg = {"type": "image", "x": 0, "y": 0, "w": round(iw, 1), "h": round(ih, 1),
          "props": {"src": "data:image/png;base64," + b64, "fit": "fill", "background": True}}
    sch = _wrap(doctype, [bg], ih)
    sch["page"]["orientation"] = "landscape" if landscape else "portrait"
    sch["__background"] = 1
    return sch


def import_format(filename, content_b64, doctype, mode="text"):
    data = base64.b64decode(content_b64.split(",")[-1])
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    if ext == "pdf":
        if mode == "image":
            return parse_pdf_image(data, doctype)
        return parse_pdf(data, doctype)
    if ext in ("xlsx", "xlsm"):
        return parse_xlsx(data, doctype)
    if ext == "docx":
        return parse_docx(data, doctype)
    if ext in ("png", "jpg", "jpeg", "webp", "bmp"):
        return parse_image(data, doctype, ext)
    frappe.throw(_("Unsupported file type: .{0}. Use PDF, XLSX, DOCX, or an image.").format(ext))
